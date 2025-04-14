import pandas as pd
import os
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime
import logging
import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def initialize_firebase():
    try:
        # 初始化 Firebase
        cred = credentials.Certificate("firebase-admin-key.json")
        firebase_admin.initialize_app(cred)
        return firestore.client()
    except Exception as e:
        logging.error(f"Firebase初始化失败: {str(e)}")
        raise

def initialize_google_sheets():
    try:
        # 使用Google Sheets API的认证范围
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        # 从环境变量或配置文件中获取Google Sheets的凭据
        creds = ServiceAccountCredentials.from_json_keyfile_name('google-sheets-credentials.json', scope)
        client = gspread.authorize(creds)
        return client
    except Exception as e:
        logging.error(f"Google Sheets初始化失败: {str(e)}")
        raise

def read_excel_with_display_values(file_path):
    try:
        # 使用openpyxl读取Excel文件以获取显示值
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # 创建一个字典来存储所有工作表的数据
        all_data = {}
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # 获取表头
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)
            
            # 获取数据
            data = []
            for row in worksheet.iter_rows(min_row=2):
                row_data = {}
                for header, cell in zip(headers, row):
                    if header is not None:  # 只处理有表头的列
                        row_data[header] = cell.value
                data.append(row_data)
            
            # 转换为DataFrame
            df = pd.DataFrame(data)
            all_data[sheet_name] = df
            
        return all_data
        
    except Exception as e:
        logging.error(f"读取Excel文件失败: {str(e)}")
        raise

def clear_collection(db, collection_name):
    try:
        # 清空指定集合
        batch_size = 500
        docs = db.collection(collection_name).limit(batch_size).stream()
        deleted = 0
        
        # 批量删除文档
        for doc in docs:
            doc.reference.delete()
            deleted += 1
            
        if deleted >= batch_size:
            # 如果删除的文档数量达到批次大小，可能还有更多文档
            return clear_collection(db, collection_name)
            
        logging.info(f"已清空集合 {collection_name}: 删除了 {deleted} 条记录")
    except Exception as e:
        logging.error(f"清空集合 {collection_name} 失败: {str(e)}")
        raise

def safe_int(value):
    try:
        if pd.isna(value) or value == '':
            return None
        if isinstance(value, str) and '+' in value:
            return int(value.replace('+', ''))
        return int(value)
    except (ValueError, TypeError):
        return None

def safe_str(value):
    if pd.isna(value):
        return ''
    return str(value).strip()

def safe_list(value):
    if pd.isna(value):
        return []
    if isinstance(value, str):
        # 如果是单个值，直接返回包含该值的列表
        return [value.strip()]
    return value if isinstance(value, list) else []

def upload_companies(db, df):
    success_count = 0
    error_count = 0
    
    for _, row in df.iterrows():
        try:
            company_id = safe_str(row['companyId'])
            if not company_id:
                logging.warning("跳过无效的公司ID")
                continue
                
            company_data = {
                'name_en': safe_str(row['name_en']),
                'name_cn': safe_str(row['name_cn']),
                'abn': safe_str(row['abn']),
                'logo': safe_str(row['logo']),
                'industry': safe_list(row['industry']),
                'website': safe_str(row['website']),
                'foundedYear': safe_int(row['foundedYear']),
                'teamSize': safe_int(row['teamSize']),
                'languages': safe_list(row['languages']),
                'shortDescription': safe_str(row['shortDescription']),
                'fullDescription': safe_str(row['fullDescription']),
                'social': safe_str(row['social']),
                'verified': bool(row['verified']) if pd.notna(row['verified']) else False,
                'createdAt': firestore.SERVER_TIMESTAMP,
                'updatedAt': firestore.SERVER_TIMESTAMP
            }
            
            # 完全覆盖现有数据
            db.collection('companies').document(company_id).set(company_data)
            success_count += 1
            logging.info(f"成功上传公司数据: {company_id}")
            
        except Exception as e:
            error_count += 1
            logging.error(f"上传公司数据失败 {company_id if 'company_id' in locals() else 'Unknown'}: {str(e)}")
            continue
            
    logging.info(f"公司数据上传完成: 成功 {success_count} 条, 失败 {error_count} 条")

def upload_offices(db, df):
    success_count = 0
    error_count = 0
    
    for _, row in df.iterrows():
        try:
            office_id = safe_str(row['officeId'])
            company_id = safe_str(row['companyId'])
            
            if not office_id or not company_id:
                logging.warning("跳过无效的办公室ID或公司ID")
                continue
                
            office_data = {
                'companyId': company_id,
                'state': safe_str(row['state']),
                'city': safe_str(row['city']),
                'address': safe_str(row['address']),
                'postalCode': safe_str(row['postalCode']),
                'contactPerson': safe_str(row['contactPerson']),
                'email': safe_str(row['email']),
                'phone': safe_str(row['phone']),
                'isHeadquarter': bool(row['isHeadquarter']) if pd.notna(row['isHeadquarter']) else False,
                'createdAt': firestore.SERVER_TIMESTAMP,
                'updatedAt': firestore.SERVER_TIMESTAMP
            }
            
            # 完全覆盖现有数据
            db.collection('offices').document(office_id).set(office_data)
            success_count += 1
            logging.info(f"成功上传办公室数据: {office_id}")
            
        except Exception as e:
            error_count += 1
            logging.error(f"上传办公室数据失败 {office_id if 'office_id' in locals() else 'Unknown'}: {str(e)}")
            continue
            
    logging.info(f"办公室数据上传完成: 成功 {success_count} 条, 失败 {error_count} 条")

def upload_services(db, df):
    success_count = 0
    error_count = 0
    
    for _, row in df.iterrows():
        try:
            service_id = safe_str(row['serviceId'])
            company_id = safe_str(row['companyId'])
            
            if not service_id or not company_id:
                logging.warning(f"跳过无效的服务ID或公司ID: serviceId={service_id}, companyId={company_id}")
                continue
                
            # 检查公司是否存在
            company_ref = db.collection('companies').document(company_id)
            if not company_ref.get().exists:
                logging.warning(f"公司不存在: {company_id}")
                continue
                
            service_data = {
                'companyId': company_id,
                'title': safe_str(row['title']),
                'description': safe_str(row['description']),
                'createdAt': firestore.SERVER_TIMESTAMP,
                'updatedAt': firestore.SERVER_TIMESTAMP
            }
            
            # 完全覆盖现有数据
            db.collection('services').document(service_id).set(service_data)
            success_count += 1
            logging.info(f"成功上传服务数据: {service_id}")
            
        except Exception as e:
            error_count += 1
            logging.error(f"上传服务数据失败 {service_id if 'service_id' in locals() else 'Unknown'}: {str(e)}")
            continue
            
    logging.info(f"服务数据上传完成: 成功 {success_count} 条, 失败 {error_count} 条")

def upload_history(db, df):
    success_count = 0
    error_count = 0
    
    for _, row in df.iterrows():
        try:
            history_id = safe_str(row['historyId'])
            company_id = safe_str(row['companyId'])
            
            if not history_id or not company_id:
                logging.warning(f"跳过无效的历史ID或公司ID: historyId={history_id}, companyId={company_id}")
                continue
                
            # 检查公司是否存在
            company_ref = db.collection('companies').document(company_id)
            if not company_ref.get().exists:
                logging.warning(f"公司不存在: {company_id}")
                continue
                
            history_data = {
                'companyId': company_id,
                'year': safe_int(row['year']),
                'event': safe_str(row['event']),
                'createdAt': firestore.SERVER_TIMESTAMP,
                'updatedAt': firestore.SERVER_TIMESTAMP
            }
            
            # 完全覆盖现有数据
            db.collection('history').document(history_id).set(history_data)
            success_count += 1
            logging.info(f"成功上传历史数据: {history_id}")
            
        except Exception as e:
            error_count += 1
            logging.error(f"上传历史数据失败 {history_id if 'history_id' in locals() else 'Unknown'}: {str(e)}")
            continue
            
    logging.info(f"历史数据上传完成: 成功 {success_count} 条, 失败 {error_count} 条")

def main():
    try:
        # 初始化 Firebase
        logging.info("正在初始化 Firebase...")
        db = initialize_firebase()
        
        # 读取Excel文件
        excel_path = "D:/IT 软件程序/qx-net-cursor-new/QX Net company data/QX Net next js.xlsx"
        logging.info(f"正在读取Excel文件: {excel_path}")
        
        # 使用新的读取方法
        all_data = read_excel_with_display_values(excel_path)
        
        # 清空现有数据
        collections = ['companies', 'offices', 'services', 'history']
        for collection in collections:
            logging.info(f"正在清空集合: {collection}")
            clear_collection(db, collection)
        
        # 上传数据
        if 'Companies' in all_data:
            upload_companies(db, all_data['Companies'])
        if 'Offices' in all_data:
            upload_offices(db, all_data['Offices'])
        if 'Services' in all_data:
            upload_services(db, all_data['Services'])
        if 'History' in all_data:
            upload_history(db, all_data['History'])
            
        logging.info("数据上传完成")
        
    except Exception as e:
        logging.error(f"数据处理失败: {str(e)}")
        raise

if __name__ == "__main__":
    main() 