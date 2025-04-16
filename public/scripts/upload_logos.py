import os
import re
import logging
import pandas as pd
import firebase_admin
from firebase_admin import credentials, storage
from datetime import timedelta
from dotenv import load_dotenv
from thefuzz import fuzz
import openpyxl
from pathlib import Path

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def initialize_firebase():
    """初始化Firebase"""
    try:
        # 加载.env文件
        load_dotenv()
        
        # 获取Firebase配置
        storage_bucket = os.getenv('FIREBASE_STORAGE_BUCKET')
        
        if not storage_bucket:
            raise ValueError("请在.env文件中设置FIREBASE_STORAGE_BUCKET")
            
        # 初始化Firebase
        if not firebase_admin._apps:
            cred = credentials.Certificate('firebase-admin-key.json')
            firebase_admin.initialize_app(cred, {
                'storageBucket': storage_bucket
            })
            
        logging.info("Firebase初始化成功")
    except Exception as e:
        logging.error(f"Firebase初始化失败: {str(e)}")
        raise

def clean_company_name(name):
    """清理公司名称，用于匹配"""
    if pd.isna(name):
        return ""
        
    name = str(name).lower()
    # 移除公司类型后缀
    name = re.sub(r'\s*(pty\.?\s*ltd\.?|p\/l|proprietary\s*limited)\.?\s*$', '', name)
    # 移除括号内容
    name = re.sub(r'\([^)]*\)', '', name)
    # 移除特殊字符
    name = re.sub(r'[&\-+]', '', name)
    # 只保留字母和数字
    name = re.sub(r'[^a-z0-9]', '', name)
    return name

def upload_image_to_firebase(file_path, use_public_url=False):
    """上传图片到Firebase Storage"""
    try:
        file_name = os.path.basename(file_path)
        bucket = storage.bucket()
        blob = bucket.blob(f"company_logos/{file_name}")  # 存储在company_logos目录下
        
        is_new_upload = False
        if not blob.exists():
            blob.upload_from_filename(file_path)
            is_new_upload = True
            logging.info(f"新上传: {file_name}")
            
        if use_public_url:
            url = f"https://storage.googleapis.com/{bucket.name}/{blob.name}"
        else:
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(days=365),  # 使用1年有效的URL
                method="GET"
            )
            
        return url, is_new_upload
    except Exception as e:
        logging.error(f"上传图片失败 {file_path}: {str(e)}")
        return None, False

def get_company_name(file_name):
    """从文件名中提取公司名称"""
    file_name = os.path.splitext(file_name)[0]
    # 移除logo、LOGO等后缀
    match = re.match(r'^(.*?)(?:logo|LOGO)', file_name, re.IGNORECASE)
    if match:
        return clean_company_name(match.group(1))
    return None

def find_best_match(company_name, df):
    """在数据框中找到最佳匹配的公司"""
    if not company_name:
        return None, 0
        
    best_match = None
    best_ratio = 0
    
    for _, row in df.iterrows():
        db_name = clean_company_name(row['name_en'])
        if not db_name:
            continue
            
        # 使用fuzz.ratio计算相似度
        ratio = fuzz.ratio(company_name, db_name)
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = row
            
    return best_match, best_ratio

def update_excel(excel_path, image_urls):
    """更新Excel文件中的logo URL"""
    try:
        # 读取Excel文件
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        # 获取列索引
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        name_col_idx = None
        logo_col_idx = None
        
        for idx, cell in enumerate(header_row, 1):
            if cell.value == 'name_en':
                name_col_idx = idx
            elif cell.value == 'logo':
                logo_col_idx = idx
                
        if not name_col_idx or not logo_col_idx:
            raise ValueError("Excel文件中缺少必要的列: name_en 或 logo")
            
        # 更新logo URL
        updated_count = 0
        for file_prefix, url in image_urls.items():
            best_match, match_ratio = find_best_match(file_prefix, pd.read_excel(excel_path))
            
            if best_match is not None and match_ratio >= 90:
                # 找到匹配的行
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
                    if row[name_col_idx-1].value == best_match['name_en']:
                        row[logo_col_idx-1].value = url
                        updated_count += 1
                        logging.info(f"更新: {best_match['name_en']} (匹配度: {match_ratio:.2f})")
                        break
            else:
                logging.warning(f"未找到匹配: {file_prefix} (最佳匹配度: {match_ratio:.2f})")
                
        # 保存Excel文件
        workbook.save(excel_path)
        logging.info(f"Excel文件更新完成，共更新了 {updated_count} 个公司的logo URL")
        
    except Exception as e:
        logging.error(f"更新Excel文件失败: {str(e)}")
        raise

def main():
    try:
        # 初始化Firebase
        initialize_firebase()
        
        # 设置路径
        current_dir = Path(__file__).parent.parent.parent
        excel_path = current_dir / "QX Net company data" / "QX Net next js.xlsx"
        logo_folder = current_dir / "public" / "logos"
        
        if not logo_folder.exists():
            logging.error(f"错误：'{logo_folder}' 文件夹不存在")
            return
            
        # 上传图片并收集URL
        image_urls = {}
        for file in logo_folder.glob("*.{png,jpg,jpeg}"):
            company_name = get_company_name(file.name)
            if company_name:
                url, is_new_upload = upload_image_to_firebase(str(file))
                if url:
                    image_urls[company_name] = url
            else:
                logging.warning(f"跳过：无法从 {file.name} 提取公司名称")
                
        # 更新Excel文件
        if image_urls:
            update_excel(str(excel_path), image_urls)
        else:
            logging.warning("没有找到可上传的logo文件")
            
    except Exception as e:
        logging.error(f"程序执行失败: {str(e)}")
        raise

if __name__ == "__main__":
    main() 